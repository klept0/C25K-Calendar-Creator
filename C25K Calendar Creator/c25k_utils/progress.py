"""
c25k_utils/progress.py
Handles progress tracking import and summary generation.
Stub: To be implemented with import and summary logic.
"""

import os
import csv
from typing import List, Dict
from openpyxl import load_workbook


def import_progress(file_path: str) -> List[Dict[str, str]]:
    """
    Import progress from an Excel (.xlsx) or CSV file. Each row should have at least
    'week', 'day', 'completed' columns. Returns a list of dicts.
    """
    progress = []
    # If a relative path is given, check in the created output folder
    if not os.path.isabs(file_path) and not os.path.exists(file_path):
        created_path = os.path.join("created", file_path)
        if os.path.exists(created_path):
            file_path = created_path
    if file_path.lower().endswith(".xlsx"):
        try:
            wb = load_workbook(file_path, data_only=True)
            # Use the first sheet or a sheet named 'Progress Tracker'
            sheet = wb.active
            if "Progress Tracker" in wb.sheetnames:
                sheet = wb["Progress Tracker"]
            headers = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {
                    headers[i]: str(row[i]) if row[i] is not None else ""
                    for i in range(len(headers))
                }
                progress.append(row_dict)
        except FileNotFoundError:
            return []
        except Exception:
            # Fallback to CSV if Excel read fails
            pass
    if not progress and file_path.lower().endswith(".csv"):
        try:
            with open(file_path, newline="", encoding="utf-8") as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    progress.append(row)
        except FileNotFoundError:
            return []
    return progress


def generate_progress_summary(progress_data: List[Dict[str, str]]) -> str:
    """
    Generate a summary of completed workouts from imported progress data.
    Expects 'completed' column to be 'yes' or 'no'.
    """
    if not progress_data:
        return "No progress data found."
    total = len(progress_data)
    completed = sum(
        1 for row in progress_data if row.get("completed", "").strip().lower() == "yes"
    )
    percent = (completed / total * 100) if total else 0
    return f"Progress: {completed}/{total} sessions completed " f"({percent:.1f}%)."
