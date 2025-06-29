"""
C25K Excel Macro Inserter

This script will insert advanced formulas/macros into the generated C25K progress tracker Excel file.
Run this after generating your tracker to auto-populate all macro columns (Current_Streak, Missed, Adjust_Plan, Milestone, Goal Progress %, etc.).

Usage:
    python3 c25k_excel_macro_inserter.py <path_to_progress_tracker.xlsx>

"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def insert_macros(filename):
    wb = load_workbook(filename)
    ws = wb[wb.sheetnames[0]]  # Assume first sheet is Progress
    max_row = ws.max_row
    # Insert Current_Streak (F)
    ws["F2"] = '=IF(D2="Y",1,0)'
    for row in range(3, max_row + 1):
        ws[f"F{row}"] = f'=IF(D{row}="Y",F{row-1}+1,0)'
    # Insert Missed (G)
    # User must update the start date (YYYY,MM,DD) as needed
    for row in range(2, max_row + 1):
        ws[f"G{row}"] = (
            f'=IF(AND(C{row}="",TODAY()-DATE(2025,7,15)+(ROW()-2)*2>2),"Missed","")'
        )
    # Insert Adjust_Plan (H) in H2 only
    ws["H2"] = (
        f'=IF(COUNTIF($D$2:$D${max_row},"N")>=3,"Consider repeating this week or shifting plan","On Track")'
    )
    # Insert Milestone (J)
    for row in range(2, max_row + 1):
        ws[f"J{row}"] = (
            f'=IF(AND(A{row}=1,B{row}=3),"First week done!",'
            f'IF(AND(A{row}=5,B{row}=3),"Halfway!",'
            f'IF(AND(A{row}=10,B{row}=3),"C25K Complete!","")))'
        )
    # Insert Goal Progress % (L2)
    ws["L2"] = f'=COUNTIF(D2:D{max_row},"Y")/COUNTA(D2:D{max_row})'
    wb.save(filename)
    print(f"Macros inserted into {filename}.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(
            "Usage: python3 c25k_excel_macro_inserter.py <path_to_progress_tracker.xlsx>"
        )
        sys.exit(1)
    insert_macros(sys.argv[1])
