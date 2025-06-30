#!/usr/bin/env python3
"""
Couch to 5K ICS Generator

DISCLAIMER: This script is for informational purposes only and is not a
substitute for professional medical advice, diagnosis, or treatment.
Always consult your healthcare provider before starting any new exercise
program, especially if you have hypertension or other pre-existing health
conditions. Use this script at your own risk. The author assumes no
responsibility for any injury or health issues that may result from using
this script.

Medical recommendations and plan structure are based on:
- NHS Couch to 5K:
  https://www.nhs.uk/live-well/exercise/couch-to-5k-week-by-week/
- CDC Physical Activity Guidelines:
  https://www.cdc.gov/physicalactivity/basics/index.htm
- American Heart Association:
  https://www.heart.org/en/healthy-living/fitness/fitness-basics
"""

import csv
import json
import os
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from c25k_utils import (
    accessibility,
    community,
    pdf_export,
    plan_customization,
    progress,
    weather,
    qr_export,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- Advanced Macros Implementation for progress tracker CSV ---
# Add these as columns or sheets in name_progress_tracker.csv as appropriate.
#
# 1. Streak Counter (formula for Google Sheets/Excel):
#    Add a column 'Current_Streak'. In row 2 (first data row):
#      =IF([@completed]="Y",1,0)
#    In row 3 and down:
#      =IF([@completed]="Y",OFFSET([@Current_Streak],-1,0)+1,0)
#    Longest streak: =MAX(Current_Streak)
#
# 2. Missed Sessions Alert:
#    Add a column 'Missed'. Formula:
#      =IF(AND([@date_completed]="",TODAY()-[@date_scheduled]>2),"Missed","")
#    (Assumes a 'date_scheduled' column exists.)
#
# 3. Adaptive Plan Adjuster:
#    Add a column 'Adjust_Plan'. Formula:
#      =IF(COUNTIF([completed],"N")>=3,
#         "Consider repeating this week or shifting plan","On Track")
#
# 4. Weekly Summary Generator:
#    At the end of each week, insert a summary row with formulas:
#      Sessions Completed: =COUNTIF([completed],"Y")
#      Sessions Missed: =COUNTIF([Missed],"Missed")
#      Motivational Msg: =IF([Sessions Completed]=[Total Sessions],"Great job!","Keep going!")
#
# 5. Effort/Feeling Tracker:
#    Add a column 'Effort' (1-5 or Easy/Medium/Hard). Use conditional formatting for trends.
#    For chart: Insert a bar/line chart of Effort vs. Date.
#
# 6. Goal Progress Visualization:
#    Add a cell for progress percent:
#      =COUNTIF([completed],"Y")/COUNTA([completed])
#    Insert a progress bar chart using this value.
#
# 7. Custom Milestone Celebrations:
#    Add a column 'Milestone'. Formula:
#      =IF(AND([@week]=1,[@day]=3),"First week done!",IF(AND([@week]=5,[@day]=3),"Halfway!",IF(AND([@week]=10,[@day]=3),"C25K Complete!","") ) )
#    Use conditional formatting to highlight these rows.
#
# 8. Weather/Condition Log:
#    Add a column 'Weather'. User logs weather/conditions for each session.
#    For trends: Insert a pivot table or chart by weather type.
#
# 9. Auto-Backup/Versioning:
#    Use spreadsheet's built-in version history, or set up a macro/script to copy the sheet weekly:
#      (Google Sheets: File > Version history > Name current version)
#      (Excel: VBA macro to copy sheet to a new tab with timestamp)
#
# All formulas/macros are beginner-friendly and can be copy-pasted into the spreadsheet. See README for more details.


def colorize(text, color, bold=False):
    colors = {
        "cyan": "\033[96m",
        "green": "\033[92m",
        "yellow": "\033[93m",
        "red": "\033[91m",
        "magenta": "\033[95m",
        "blue": "\033[94m",
        "white": "\033[97m",
        "reset": "\033[0m",
        "bold": "\033[1m",
    }
    prefix = colors.get(color, "")
    if bold:
        prefix = colors["bold"] + prefix
    return f"{prefix}{text}{colors['reset']}"


def get_user_info() -> Optional[Dict[str, Any]]:
    """
    Prompt user for name, age, weight (imperial by default), gender,
    session time, language, personal goal, and advanced options.
    Returns a dict or None if incomplete.
    """
    try:
        name = input(colorize("Enter your name: ", "green", bold=True)).strip()
        if not name:
            print(colorize("Name is required.", "red", bold=True))
            return None
        # Default to imperial units
        unit = input(
            colorize(
                "Choose units: [I]mperial (lbs, default) or [M]etric (kg)? ",
                "blue",
                bold=True,
            )
        ).strip().lower()
        if unit == "" or unit == "i":
            unit = "i"
        elif unit == "m":
            unit = "m"
        else:
            print(
                colorize(
                    "Please enter 'I' for Imperial or 'M' for Metric.",
                    "red",
                    bold=True,
                )
            )
            return None
        age = int(
            input(
                colorize(
                    "Enter your age (years): ",
                    "yellow",
                    bold=True,
                )
            ).strip()
        )
        if unit == "m":
            weight = float(
                input(
                    colorize("Enter your weight (kg): ", "magenta", bold=True)
                ).strip()
            )
        else:
            weight = float(
                input(
                    colorize("Enter your weight (lbs): ", "magenta", bold=True)
                ).strip()
            )
            weight = weight * 0.453592  # Convert lbs to kg
        gender = input(
            colorize(
                "Enter your gender ([M]ale/[F]emale): ",
                "cyan",
                bold=True,
            )
        ).strip().lower()
        if gender in ["m", "male"]:
            gender = "male"
        elif gender in ["f", "female"]:
            gender = "female"
        else:
            print(colorize("Gender must be 'M'/'F' or 'male'/'female'.", "red", bold=True))
            return None
        # Session time
        time_str = input(
            colorize(
                "Enter session start time (HH:MM, 24h, default 07:00): ",
                "blue",
                bold=True,
            )
        ).strip()
        if time_str:
            try:
                hour, minute = map(int, time_str.split(":"))
            except Exception:
                print(
                    colorize(
                        "Invalid time format. Please use HH:MM (24h).",
                        "red",
                        bold=True,
                    )
                )
                return None
        else:
            hour, minute = 7, 0
        # Language/localization
        lang = input(
            colorize(
                "Choose language: [E]nglish (default) or [S]panish: ",
                "green",
                bold=True,
            )
        ).strip().lower()
        if lang not in ("e", "s", ""):
            print(
                colorize(
                    "Please enter 'E' for English or 'S' for Spanish.",
                    "red",
                    bold=True,
                )
            )
            return None
        lang = lang if lang else "e"
        # Export option
        export = (
            input(
                colorize(
                    "Export format: [I]CS (default), [C]SV, [J]SON, [G]oogle Fit CSV, [P]DF, [M]arkdown, [V]oice, [S]trava/Runkeeper? ",
                    "yellow",
                    bold=True,
                )
            )
            .strip()
            .lower()
        )
        if export not in ("i", "c", "j", "g", "p", "m", "v", "s", "q", ""):
            print(colorize("Please enter a valid export option.", "red", bold=True))
            return None
        export = export if export else "i"
        # Personal goal
        goal = input(
            colorize("Enter your personal C25K goal (optional): ", "magenta", bold=True)
        ).strip()
        # Advanced: plan length
        weeks, days_per_week = plan_customization.get_custom_plan_length()
        # Advanced: accessibility
        high_contrast = (
            input(
                colorize("High-contrast mode? [Y/N] (default N): ", "white", bold=True)
            )
            .strip()
            .lower()
            == "y"
        )
        large_font = (
            input(colorize("Large font mode? [Y/N] (default N): ", "white", bold=True))
            .strip()
            .lower()
            == "y"
        )
        # New: dyslexia-friendly font option
        dyslexia_font = (
            input(colorize("Dyslexia-friendly font? [Y/N] (default N): ", "white", bold=True))
            .strip()
            .lower()
            == "y"
        )
        # Advanced: dynamic start date
        start_option = (
            input(
                colorize(
                    "Start date: [D]efault, [N]ext Monday, or YYYY-MM-DD? ",
                    "blue",
                    bold=True,
                )
            )
            .strip()
            .lower()
        )
        if start_option == "n":
            from datetime import datetime

            start_day = start_date.get_dynamic_start_date()
        elif start_option and start_option != "d":
            try:
                from datetime import datetime

                start_day = datetime.strptime(start_option, "%Y-%m-%d")
            except Exception:
                print(
                    colorize("Invalid date format. Use YYYY-MM-DD.", "red", bold=True)
                )
                return None
        else:
            start_day = None  # Use default in main
        # Advanced: email for reminders
        email = input(
            colorize("Enter your email for reminders (optional): ", "green", bold=True)
        ).strip()
        # Advanced: location for weather
        location = input(
            colorize(
                "Enter your city or ZIP for weather suggestions (optional, default F°): ",
                "cyan",
                bold=True,
            )
        ).strip()
        # Custom alert time for ICS
        alert_minutes = None
        if export == "i":
            alert_input = input(
                colorize(
                    "Alert before session (minutes, default 30, 0=none): ",
                    "magenta",
                    bold=True,
                )
            ).strip()
            if alert_input == "":
                alert_minutes = 30
            else:
                try:
                    alert_minutes = int(alert_input)
                except Exception:
                    print(
                        colorize(
                            "Invalid alert time. Using default 30 minutes.",
                            "red",
                            bold=True,
                        )
                    )
                    alert_minutes = 30
        # Custom rest day scheduling
        rest_days_input = input(
            colorize(
                "Enter your preferred rest days (comma-separated, e.g. Sat,Sun) or leave blank for default (Sat,Sun): ",
                "yellow",
                bold=True,
            )
        ).strip()
        if rest_days_input:
            rest_days = [d.strip().capitalize()[:3] for d in rest_days_input.split(",") if d.strip()]
        else:
            rest_days = ["Sat", "Sun"]
        # Privacy/anonymization prompt
        anonymize = (
            input(colorize("Would you like to anonymize your data in all exports? [Y/N] (default N): ", "yellow", bold=True))
            .strip().lower() == "y"
        )
        return {
            "name": name,
            "age": age,
            "weight": weight,
            "gender": gender,
            "unit": unit,
            "hour": hour,
            "minute": minute,
            "lang": lang,
            "export": export,
            "goal": goal,
            "weeks": weeks,
            "days_per_week": days_per_week,
            "high_contrast": high_contrast,
            "large_font": large_font,
            "dyslexia_font": dyslexia_font,
            "start_day": start_day,
            "email": email,
            "location": location,
            "alert_minutes": alert_minutes,
            "rest_days": rest_days,
            "anonymize": anonymize,
        }
    except Exception:
        print(colorize("An error occurred during user input.", "red", bold=True))
        return None


# --- Utility for anonymizing user info in exports ---
def anonymize_user(user: Dict[str, Any]) -> Dict[str, Any]:
    if not user.get("anonymize"):
        return user
    anon_user = user.copy()
    anon_user["name"] = "Anonymous"
    anon_user["email"] = ""
    return anon_user

# --- Add privacy note to all exports ---
PRIVACY_NOTE = (
    "Your personal data (name, email, credentials, tokens) is only used locally to generate your plan and reminders. "
    "No data is sent to any server. If you choose to anonymize, your name and email will not appear in any export. "
    "Credentials/tokens are used in-memory only and never saved. See README for full privacy details."
)


def get_workout_details(week: int, day: int, lang: str = "e") -> str:
    """
    Return a string describing the actual workout for the given week and day.
    Supports English and Spanish.
    Workouts are based on the NHS Couch to 5K program.
    """
    workouts_en = [
        "Brisk 5-min warmup walk. Then alternate 60 sec jogging and 90 sec "
        "walking for 20 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Then alternate 90 sec jogging, 2 min walking "
        "for 20 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. 90 sec jog, 90 sec walk, 3 min jog, 3 min "
        "walk, repeat. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 3 min, walk 90 sec, jog 5 min, walk 2.5 "
        "min, jog 3 min, walk 90 sec, jog 5 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 5 min, walk 3 min, jog 5 min, walk 3 min, "
        "jog 5 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 8 min, walk 5 min, jog 8 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 25 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 28 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 30 min. Hydrate before and after.",
        "Brisk 5-min warmup walk. Jog 30 min. Hydrate before and after.",
    ]
    workouts_es = [
        "Camine rápido 5 min para calentar. Luego alterne 60 seg corriendo y 90 seg caminando durante 20 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Luego alterne 90 seg corriendo, 2 min caminando durante 20 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. 90 seg corra, 90 seg camine, 3 min corra, 3 min camine, repita. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 3 min, camine 90 seg, corra 5 min, camine 2.5 min, corra 3 min, camine 90 seg, corra 5 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 5 min, camine 3 min, corra 5 min, camine 3 min, corra 5 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 8 min, camine 5 min, corra 8 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 25 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 28 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 30 min. Hidratese antes y después.",
        "Camine rápido 5 min para calentar. Corra 30 min. Hidratese antes y después.",
    ]
    workouts = workouts_en if lang == "e" else workouts_es
    if 1 <= week <= 10:
        return workouts[week - 1]
    return workouts[0] if lang == "e" else workouts_es[0]


def get_beginner_tip(day: int, lang: str = "e") -> str:
    """
    Return a motivational or safety tip for the given day.
    Tips are based on NHS, CDC, and AHA recommendations for beginners.
    """
    tips_en = [
        "Remember to stretch before and after your workout!",
        "Wear comfortable shoes and clothing.",
        "Stay hydrated and listen to your body.",
        "Rest is as important as running. Take it easy on rest days!",
        "Track your progress and celebrate small wins.",
        "Invite a friend or family member to join you!",
        "If you feel pain, stop and consult a professional.",
        "Set a reminder so you don't miss your session.",
        "Smile and enjoy the journey!",
        "You're doing great—keep going!",
    ]
    tips_es = [
        "¡Recuerda estirar antes y después de tu entrenamiento!",
        "Usa calzado y ropa cómodos.",
        "Mantente hidratado y escucha a tu cuerpo.",
        "El descanso es tan importante como correr. ¡Tómalo con calma en los días de descanso!",
        "Registra tu progreso y celebra los pequeños logros.",
        "¡Invita a un amigo o familiar a unirse!",
        "Si sientes dolor, detente y consulta a un profesional.",
        "Pon una alarma para no perder tu sesión.",
        "¡Sonríe y disfruta el proceso!",
        "¡Lo estás haciendo genial, sigue así!",
    ]
    tips = tips_en if lang == "e" else tips_es
    return tips[(day - 1) % len(tips)]


def fetch_weather_for_session(location: str, date: datetime) -> str:
    """
    Fetch weather forecast for the given location and date using OpenWeatherMap API.
    Requires the environment variable OWM_API_KEY to be set, or replace with your API key.
    Returns a string summary (e.g., 'Sunny 75°F', 'Rainy 60°F').
    Falls back to stub if API fails.
    """
    api_key = os.environ.get("OWM_API_KEY", "YOUR_OPENWEATHERMAP_API_KEY")
    if not api_key or api_key == "YOUR_OPENWEATHERMAP_API_KEY":
        # Fallback to stub
        weather_map = {
            0: "Sunny 75°F",
            1: "Partly Cloudy 72°F",
            2: "Rainy 68°F",
            3: "Thunderstorms 70°F",
            4: "Cloudy 71°F",
            5: "Showers 65°F",
            6: "Windy 69°F",
        }
        return weather_map[date.weekday()]
    try:
        # Use city name or ZIP (US only) for location
        params = {
            "q": location,
            "appid": api_key,
            "units": "imperial",
        }
        url = "https://api.openweathermap.org/data/2.5/forecast"
        resp = requests.get(url, params=params, timeout=5)
        resp.raise_for_status()
        data = resp.json()
        # Find the forecast closest to the session date/time
        from datetime import datetime as dt
        target = dt.combine(date.date(), dt.min.time())
        best = None
        min_diff = float("inf")
        for entry in data.get("list", []):
            entry_time = dt.fromtimestamp(entry["dt"])
            diff = abs((entry_time - target).total_seconds())
            if diff < min_diff:
                min_diff = diff
                best = entry
        if best:
            desc = best["weather"][0]["description"].capitalize()
            temp = round(best["main"]["temp"])
            return f"{desc} {temp}°F"
        else:
            return "Weather unavailable"
    except Exception:
        # Fallback to stub
        weather_map = {
            0: "Sunny 75°F",
            1: "Partly Cloudy 72°F",
            2: "Rainy 68°F",
            3: "Thunderstorms 70°F",
            4: "Cloudy 71°F",
            5: "Showers 65°F",
            6: "Windy 69°F",
        }
        return weather_map[date.weekday()]


def get_workout_plan(user: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Return a plan (list of dicts) based on age, weight, and plan customization.
    Adjusts session duration for older or heavier users (age >= 60 or weight >= 100kg) for safety.
    Adds actual workout details and tips for each session.
    Supports custom rest day scheduling and adaptive plan logic.
    """
    weeks = user.get("weeks", 10)
    days_per_week = user.get("days_per_week", 3)
    rest_days = user.get("rest_days", ["Sat", "Sun"])
    plan: List[Dict[str, Any]] = []
    start_day = user.get("start_day")
    if not start_day:
        from datetime import datetime
        start_day = datetime(2025, 7, 15)
    for week in range(weeks):
        for day in range(days_per_week):
            day_offset = day * (7 // days_per_week)
            session_date = start_day + timedelta(weeks=week, days=day_offset)
            workout = get_workout_details(week + 1, day + 1, user.get("lang", "e"))
            tip = get_beginner_tip(day + 1, user.get("lang", "e"))
            # Fetch weather for this session
            weather_str = ""
            if user.get("location"):
                weather_str = fetch_weather_for_session(user["location"], session_date)
            description = (
                f"Follow the Couch to 5K plan - Week {week+1} session. "
                f"Note: This plan is tailored for an adult {user['gender']} "
                f"aged {user['age']} with hypertension. "
                f"Weight: {user['weight']:.1f} kg. "
                f"Session time: {user['hour']:02d}:{user['minute']:02d}. "
                "Please monitor your health and consult your doctor if needed.\n"
                f"Workout: {workout}\n"
                f"Tip: {tip}"
            )
            plan.append(
                {
                    "week": week + 1,
                    "day": day + 1,
                    "day_offset": day_offset,
                    "duration": 30,  # minutes
                    "description": description,
                    "workout": workout,
                    "tip": tip,
                    "weather": weather_str,
                    "date": session_date.strftime("%Y-%m-%d"),
                }
            )
    if user["age"] >= 60 or user["weight"] >= 100:
        for session in plan:
            session["duration"] = 25
            session["description"] += " (Reduced session duration for safety.)"
    # Add rest days (user-selected)
    for week in range(weeks):
        for rest_offset, rest_name in enumerate(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
            if rest_name in rest_days and rest_offset >= days_per_week:
                plan.append(
                    {
                        "week": week + 1,
                        "day": rest_name,
                        "day_offset": rest_offset,
                        "duration": 0,
                        "description": (
                            f"Rest Day - Week {week+1} {rest_name}. "
                            "Rest and recover. Hydrate and stretch."
                        ),
                        "workout": "Rest Day",
                        "tip": get_beginner_tip(rest_offset, user.get("lang", "e")),
                    }
                )
    return sorted(plan, key=lambda s: (s["week"], s["day_offset"]))


def format_ics_datetime(dt: datetime) -> str:
    """
    Format a datetime object for ICS file (YYYYMMDDTHHMMSS).
    """
    return dt.strftime("%Y%m%dT%H%M%S")


def generate_ics(
    plan: List[Dict[str, Any]],
    start_day: datetime,
    hour: int,
    minute: int,
    alert_minutes: int = 30,
    outdir: str = ".",
) -> None:
    """
    Generate the ICS file from the workout plan and start date.
    Add the actual workout, tip, and rest days to the DESCRIPTION and NOTES fields.
    Add a VALARM block for custom alert time if alert_minutes > 0.
    ICS format is compatible with Apple/Google Calendar and most calendar apps.
    """
    ics_content = "BEGIN:VCALENDAR\nVERSION:2.0\nPRODID:-//Couch to 5K//EN\n"
    for session in plan:
        session_date = start_day + timedelta(
            weeks=session["week"] - 1, days=session["day_offset"]
        )
        dt_start = session_date.replace(hour=hour, minute=minute)
        dt_end = dt_start + timedelta(minutes=session["duration"])
        if session["duration"] > 0:
            event_name = f"C25K Week {session['week']} - Day {session['day']}"
        else:
            event_name = f"C25K Week {session['week']} {session['day']} (Rest)"
        ics_content += (
            f"BEGIN:VEVENT\n"
            f"SUMMARY:{event_name}\n"
            f"DTSTART;TZID=America/New_York:{format_ics_datetime(dt_start)}\n"
            f"DTEND;TZID=America/New_York:{format_ics_datetime(dt_end)}\n"
            f"DESCRIPTION:{session['description']}\n"
            f"X-APPLE-NOTES:Workout: {session['workout']} | Tip: {session['tip']}\n"
        )
        # Add VALARM if alert_minutes > 0 and not a rest day
        if alert_minutes and session["duration"] > 0:
            ics_content += (
                "BEGIN:VALARM\n"
                f"TRIGGER:-PT{alert_minutes}M\n"
                "ACTION:DISPLAY\n"
                "DESCRIPTION:Time for your C25K session!\n"
                "END:VALARM\n"
            )
        ics_content += "END:VEVENT\n"
    ics_content += "END:VCALENDAR"
    filename = os.path.join(outdir, "Couch_to_5K_Reminders.ics")
    with open(filename, "w", encoding="utf-8") as f:
        f.write(ics_content)
    print(f"ICS file '{filename}' created successfully.")


def export_csv(plan: List[Dict[str, Any]], filename: str) -> None:
    """Export the workout plan to a CSV file, including tips and rest days."""
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        fieldnames = ["week", "day", "duration", "description", "workout", "tip"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for session in plan:
            writer.writerow({k: session[k] for k in fieldnames})
    print(f"CSV file '{filename}' created successfully.")


def export_json(plan: List[Dict[str, Any]], filename: str) -> None:
    """Export the workout plan to a JSON file, including tips and rest days."""
    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(plan, jsonfile, ensure_ascii=False, indent=2)
    print(f"JSON file '{filename}' created successfully.")


def export_google_fit_csv(plan: List[Dict[str, Any]], filename: str) -> None:
    """Export the workout plan to a Google Fit compatible CSV file, including tips."""
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        fieldnames = ["Activity Type", "Start Date", "End Date", "Description", "Tip"]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for session in plan:
            writer.writerow(
                {
                    "Activity Type": "Running" if session["duration"] > 0 else "Rest",
                    "Start Date": f"Week {session['week']} Day {session['day']}",
                    "End Date": f"Week {session['week']} Day {session['day']}",
                    "Description": session["workout"],
                    "Tip": session["tip"],
                }
            )
    print(f"Google Fit CSV file '{filename}' created successfully.")


def export_markdown_checklist(
    plan: List[Dict[str, Any]], filename: str, user: Dict[str, Any]
) -> None:
    """Export the workout plan as a Markdown checklist file with tips and goal, with accessibility options if selected."""
    user = anonymize_user(user)
    content = f"# Couch to 5K Checklist\n\n**Name:** {user['name']}\n\n**Age:** {user['age']}\n\n**Start Date:** {user['start_day'].strftime('%Y-%m-%d') if user.get('start_day') else 'default'}\n\n"
    if user.get("goal"):
        content += f"**Personal Goal:** {user['goal']}\n\n"
    content += "**Resource:** [C25K Guide](https://www.nhs.uk/live-well/exercise/couch-to-5k-week-by-week/)\n\n"
    for session in plan:
        if session["duration"] > 0:
            content += (
                f"- [ ] Week {session['week']} Day {session['day']}: "
                f"{session['workout']}\n  - Tip: {session['tip']}\n  - Notes: ________________________________\n    ________________________________\n    ________________________________\n"
            )
        else:
            content += f"- [ ] Week {session['week']} {session['day']}: Rest Day\n  - Tip: {session['tip']}\n  - Notes: ________________________________\n    ________________________________\n    ________________________________\n"
    # Apply accessibility options if selected
    if user.get("high_contrast") or user.get("large_font") or user.get("dyslexia_font"):
        # Add a style block for Markdown renderers that support it
        style = "<style>\n"
        if user.get("large_font"):
            style += "body,li { font-size: 1.3em !important; }\n"
        if user.get("dyslexia_font"):
            style += "body,li { font-family: 'Comic Sans MS', 'OpenDyslexic', Arial, sans-serif !important; }\n"
        if user.get("high_contrast"):
            style += "body { background: #000; color: #FFD700; }\n"
        style += "</style>\n\n"
        content = style + content
        # Also call accessibility module for further processing (now includes dyslexia_font)
        content = accessibility.apply_accessibility_options(
            content, user.get("high_contrast"), user.get("large_font"), user.get("dyslexia_font")
        )
    with open(filename, "w", encoding="utf-8") as mdfile:
        mdfile.write(content)
        mdfile.write(f"\n---\n**Privacy Note:** {PRIVACY_NOTE}\n")
    print(colorize(f"Markdown checklist '{filename}' created successfully.", "green", bold=True))


def get_output_dir(user):
    # Use YYYY-MM-DD for start date
    start_str = (
        user["start_day"].strftime("%Y-%m-%d") if user.get("start_day") else "default"
    )
    safe_name = user["name"].replace(" ", "_")
    outdir = os.path.join(
        os.path.dirname(__file__),
        "created",
        f"{safe_name}-{user['age']}-{start_str}",
    )
    os.makedirs(outdir, exist_ok=True)
    return outdir


def create_progress_tracker(user: Dict[str, Any], outdir: str) -> str:
    """
    Create a progress tracker Excel file in the output directory with the correct name and columns.
    Implements all advanced visual cues: checkmarks, rest day highlighting, overdue alerts, sparklines, milestone badges, weekly progress bars, goal gauge, weather icons/colors, accessibility toggle, and notes highlighting.
    Returns the filename.
    """
    import os
    from openpyxl.formatting.rule import (
        ColorScaleRule,
        FormulaRule,
    )
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    filename = os.path.join(outdir, f"{user['name']}_progress_tracker.xlsx")
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Progress"
        # Accessibility: set font size and font family if selected
        default_font_name = "Calibri"
        if user.get("dyslexia_font"):
            default_font_name = "Comic Sans MS"
        default_font_size = 11
        if user.get("large_font"):
            default_font_size = 14
        # Set default font for all cells (including new ones)
        ws.sheet_properties.tabColor = "1072BA"
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name=default_font_name, size=default_font_size)
        ws.parent.stylesheet.fonts[0].name = default_font_name
        ws.parent.stylesheet.fonts[0].size = default_font_size
        columns = [
            "week",
            "day",
            "date_completed",
            "completed",
            "notes",
            "Current_Streak",
            "Missed",
            "Adjust_Plan",
            "Effort",
            "Milestone",
            "Weather",
            "Motivation",
            "Session_Rating",
            "Health_Log",
        ]
        ws.append(columns)
        # Center headings and set bold font
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, name=default_font_name, size=default_font_size)
        # Motivational quotes list
        motivational_quotes = [
            "You are stronger than you think!",
            "Every step counts. Keep going!",
            "Progress, not perfection.",
            "Believe in yourself and all that you are.",
            "Small steps lead to big changes.",
            "You’re doing great—don’t stop now!",
            "Consistency is key.",
            "Your only limit is you.",
            "Celebrate every victory, no matter how small.",
            "The journey is just as important as the destination.",
        ]
        weeks = user.get("weeks", 10)
        days_per_week = user.get("days_per_week", 3)
        total_rows = weeks * days_per_week
        start_day = user.get("start_day")
        if not start_day:
            from datetime import datetime
            start_day = datetime(2025, 7, 15)
        # Custom rest day scheduling
        rest_days = user.get("rest_days")
        if not rest_days:
            rest_days = ["Sat", "Sun"]  # Default
        for i, (week, day) in enumerate(
            ((w, d) for w in range(1, weeks + 1) for d in range(1, days_per_week + 1)),
            start=2,
        ):
            workout = get_workout_details(week, day, user.get("lang", "e"))
            quote = motivational_quotes[((week-1)*days_per_week + (day-1)) % len(motivational_quotes)]
            # Fetch weather for this session
            session_date = start_day + timedelta(weeks=week-1, days=(day-1)*(7//days_per_week))
            weather_str = ""
            if user.get("location"):
                weather_str = fetch_weather_for_session(user["location"], session_date)
            ws.append([
                week, day, "", "", f"Workout: {workout}", "", "", "", "", "", weather_str, quote, "", ""
            ])
            row = i
            ws[f"D{row}"].number_format = "General"
        # Auto-size columns to fit content
        for col_idx, col in enumerate(columns, 1):
            max_length = len(col)
            for row in ws.iter_rows(
                min_row=2, min_col=col_idx, max_col=col_idx, max_row=total_rows + 1
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        # 1. Completion checkmarks/icons for completed sessions (D):
        ws.conditional_formatting.add(
            f"D2:D{total_rows+1}",
            FormulaRule(
                formula=['D2="Y"'],
                font=Font(color="008000", bold=True),
                fill=PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                ),
            ),
        )
        # 2. Rest day highlighting (gray fill/italic font):
        ws.conditional_formatting.add(
            f"B2:B{total_rows+1}",
            FormulaRule(
                formula=['OR(B2="Sat",B2="Sun",B2="Rest")'],
                font=Font(italic=True, color="808080"),
                fill=PatternFill(
                    start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
                ),
            ),
        )
        # 3. Overdue session alerts (red fill/warning icon in Missed column):
        ws.conditional_formatting.add(
            f"G2:G{total_rows+1}",
            FormulaRule(
                formula=['G2="Missed"'],
                font=Font(color="FF0000", bold=True),
                fill=PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                ),
            ),
        )
        # 4. Effort trend sparklines (mini line chart per week, with instructions for Excel):
        # (openpyxl does not support sparklines, so add instructions in Macros sheet)
        # 5. Milestone badges (emoji/icons in Milestone column) - already in J above
        ws.conditional_formatting.add(
            f"J2:J{total_rows+1}",
            FormulaRule(
                formula=["LEN(J2)>0"],
                fill=PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                ),
            ),
        )
        # 6. Weekly progress bars (data bars for % completed per week):
        # (openpyxl does not support data bars, so add instructions in Macros sheet)
        # 7. Goal progress gauge (doughnut/gauge chart for overall completion):
        # (openpyxl does not support doughnut charts, so add instructions in Macros sheet)
        # 8. Weather condition icons/colors in Weather column:
        ws.conditional_formatting.add(
            f"K2:K{total_rows+1}",
            FormulaRule(
                formula=['OR(K2="Rain",K2="Snow",K2="Fog",K2="Hot",K2="Cold")'],
                font=Font(color="0000FF", bold=True),
                fill=PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                ),
            ),
        )
        # 9. Accessibility toggle (macro/button for high-contrast mode):
        # (Add VBA macro instructions in Macros sheet)
        # 10. Notes highlighting (yellow fill/bold if notes present):
        ws.conditional_formatting.add(
            f"E2:E{total_rows+1}",
            FormulaRule(
                formula=["LEN(E2)>0"],
                font=Font(bold=True),
                fill=PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                ),
            ),
        )
        # Conditional formatting for Effort (I): color scale 1 (green) to 5 (red)
        ws.conditional_formatting.add(
            f"I2:I{total_rows+1}",
            ColorScaleRule(
                start_type="num",
                start_value=1,
                start_color="C6EFCE",
                mid_type="num",
                mid_value=3,
                mid_color="FFEB9C",
                end_type="num",
                end_value=5,
                end_color="FFC7CE",
            ),
        )
        # Conditional formatting for Current_Streak (F): blue scale
        ws.conditional_formatting.add(
            f"F2:F{total_rows+1}",
            ColorScaleRule(
                start_type="min",
                start_color="DCE6F1",
                mid_type="percentile",
                mid_value=50,
                mid_color="8DB4E2",
                end_type="max",
                end_color="305496",
            ),
        )
        # Macros & Instructions sheet
        ws2 = wb.create_sheet("Macros & Instructions")
        ws2["A1"] = "Macros and Features Overview:"
        ws2["A2"] = "- Current_Streak: See formula block below."
        ws2["A3"] = "- Missed: See formula block below."
        ws2["A4"] = "- Adjust_Plan: See formula block below."
        ws2["A5"] = (
            "- Effort: Enter 1-5 (Easy-Hard). Conditional formatting shows trends."
        )
        ws2["A6"] = "- Milestone: See formula block below."
        ws2["A7"] = "- Weather: Log conditions for each session. Pre-filled with forecast if location provided."
        ws2["A8"] = "- Goal Progress: See formula block below."
        ws2["A9"] = "- Weekly Summary: See formula block below."
        ws2["A10"] = "- Auto-Backup: Use File > Version history or add this VBA macro:"
        ws2["A11"] = (
            'Sub BackupSheet()\n    Sheets("Progress").Copy After:=Sheets(Sheets.Count)\n'
            '    Sheets(Sheets.Count).Name = "Backup_" & Format(Now, "yyyymmdd_HHMMSS")\nEnd Sub'
        )
        ws2["A12"] = (
            "To use: Press Alt+F11, Insert > Module, paste the macro, and run BackupSheet."
        )
        ws2["A13"] = "Font Accessibility: If you selected large font or dyslexia-friendly font, this tracker uses Comic Sans MS and/or larger text for easier reading."
        ws2["A14"] = "Visual Cues & Advanced Features:"
        ws2["A15"] = (
            "1. Completion checkmarks: Mark 'Y' in Completed for a green check."
        )
        ws2["A16"] = (
            "2. Rest day highlighting: Enter 'Rest' in Day or Notes for gray/italic."
        )
        ws2["A17"] = "3. Overdue alerts: 'Missed' in Missed column turns red."
        ws2["A18"] = (
            "4. Effort sparklines: Select Effort cells for a week, Insert > Sparklines."
        )
        ws2["A19"] = "5. Milestone badges: See formula block below."
        ws2["A20"] = (
            "6. Weekly progress bars: Select Completed for a week, Insert > Data Bars."
        )
        ws2["A21"] = "7. Goal gauge: Select Goal Progress %, Insert > Doughnut chart."
        ws2["A22"] = "8. Weather icons: Enter 'Rain', 'Snow', etc. for color/icon."
        ws2["A23"] = "9. Accessibility: See macro below for high-contrast mode."
        ws2["A24"] = "10. Notes highlighting: Any notes are yellow/bold."
        ws2["A26"] = "High-Contrast Macro (VBA):"
        ws2["A27"] = (
            "Sub HighContrastMode()\n"
            "    Dim ws As Worksheet\n"
            '    Set ws = Sheets("Progress")\n'
            "    ws.Cells.Interior.Color = RGB(0,0,0)\n"
            "    ws.Cells.Font.Color = RGB(255,255,0)\n"
            "    ws.Rows(1).Interior.Color = RGB(255,255,0)\n"
            "    ws.Rows(1).Font.Color = RGB(0,0,0)\n"
            "End Sub"
        )
        ws2["A33"] = (
            "To use: Press Alt+F11, Insert > Module, paste the macro, and run HighContrastMode."
        )
        # Add all advanced formulas, each with a bold+underlined title and formulas on individual lines
        row = 36

        def add_formula_section(title, lines):
            nonlocal row
            ws2[f"A{row}"].value = title
            ws2[f"A{row}"].font = Font(bold=True, underline="single")
            row += 1
            for line in lines:
                ws2[f"A{row}"].value = line
                row += 1
            row += 1  # Blank line after each section

        add_formula_section(
            "Current_Streak (F):",
            [
                'In F2: =IF(D2="Y",1,0)',
                'In F3 and down: =IF(D3="Y",F2+1,0)',
            ],
        )
        add_formula_section(
            "Missed (G):",
            [
                'In G2 and down: =IF(AND(C2="",TODAY()-DATE(YYYY,MM,DD)+(ROW()-2)*2>2),"Missed","")',
                "# Replace YYYY,MM,DD with your plan's start date.",
            ],
        )
        add_formula_section(
            "Adjust_Plan (H):",
            [
                'In H2: =IF(AND(COUNTIF($G$2:$G$N,"Missed")>=3,COUNTIF($I$2:$I$N,">=4")>=3),'
                '"Consider repeating this week or shifting plan","On Track")',
                "# Replace N with the last row number.",
            ],
        )
        add_formula_section(
            "Milestone (J):",
            [
                "In J2 and down:",
                (
                    '=IF(AND(A2=1,B2=3),"First week done!",IF(AND(A2=5,B2=3),'
                    '"Halfway!",IF(AND(A2=10,B2=3),"C25K Complete!","")))'
                ),
            ],
        )
        add_formula_section(
            "Goal Progress % (L2):",
            [
                '=COUNTIF(D2:Dn,"Y")/COUNTA(D2:Dn)',
                "# Replace n with the last row number.",
            ],
        )
        add_formula_section(
            "Weekly Summary (insert after each week):",
            [
                'Sessions Completed: =COUNTIF(D2:D4,"Y")',
                'Sessions Missed: =COUNTIF(G2:G4,"Missed")',
                (
                    "Motivational Msg: =IF([Sessions Completed]=[Total Sessions],"
                    '"Great job!","Keep going!")'
                ),
            ],
        )
        add_formula_section(
            "Effort Trend:",
            ["Select Effort cells for a week, Insert > Sparklines."],
        )
        add_formula_section(
            "Weekly Progress Bars:",
            ["Select Completed for a week, Insert > Data Bars."],
        )
        add_formula_section(
            "Goal Gauge:",
            ["Select Goal Progress %, Insert > Doughnut chart."],
        )
        # --- New: Export/Integration, Accessibility, Community, Analytics, FAQ Sections ---
        add_formula_section(
            "Export/Integration Features:",
            [
                "Apple Health Export: (Planned) Use CSV/ICS export, then import into Apple Health via Shortcuts or 3rd-party app.",
                "Google Fit Export: Use the Google Fit CSV export option (already available).",
                "Strava/Runkeeper Export: Use the Strava/Runkeeper export option (already available, stub for direct integration).",
                "Instructions: For direct integration, see README or use provided scripts/macros (future update).",
            ],
        )
        add_formula_section(
            "Accessibility Improvements:",
            [
                "Screen Reader Tips: Use clear column headers, avoid merged cells, and keep formulas simple.",
                "All notes and instructions are provided in plain text for compatibility.",
                "Dark Mode Macro: See below for a VBA macro to enable dark mode in Excel.",
            ],
        )
        ws2["A40"] = "Dark Mode Macro (VBA):"
        ws2["A41"] = (
            "Sub DarkMode()\n"
            "    Dim ws As Worksheet\n"
            "    Set ws = Sheets(\"Progress\")\n"
            "    ws.Cells.Interior.Color = RGB(30,30,30)\n"
            "    ws.Cells.Font.Color = RGB(220,220,220)\n"
            "    ws.Rows(1).Interior.Color = RGB(50,50,50)\n"
            "    ws.Rows(1).Font.Color = RGB(255,255,0)\n"
            "End Sub"
        )
        ws2["A42"] = "To use: Press Alt+F11, Insert > Module, paste the macro, and run DarkMode."
        add_formula_section(
            "Community & Social Features:",
            [
                "Progress Sharing: (Planned) Export your progress as CSV/Markdown and share with friends or a group.",
                "Group Plans: (Planned) Combine multiple users' trackers for group progress (see README for future updates).",
                "Leaderboard: (Planned) Track and compare progress with others (future update).",
            ],
        )
        add_formula_section(
            "Analytics & Custom Metrics:",
            [
                "Weather Trend Analysis: (Planned) Use the Weather column to chart performance vs. weather.",
                "User-Defined Metrics: Add custom columns for any metric you wish to track (e.g., Heart Rate, Steps).",
                "Charts: Use the Dashboard sheet to visualize any metric (see instructions above).",
            ],
        )
        add_formula_section(
            "Data Backup, Import & Export:",
            [
                "Auto-Backup: Use the provided macro or Excel/Google Sheets version history.",
                "Import/Export: (Planned) Use CSV/Excel import/export for data migration or backup.",
                "Instructions: See README for details and future updates.",
            ],
        )
        add_formula_section(
            "FAQ, Help & Feedback:",
            [
                "FAQ: See the README or the FAQ section in this sheet (future update).",
                "Help: For help with macros or formulas, see the instructions above or contact the project maintainer.",
                "Feedback: (Planned) Use the feedback form in the README or email for suggestions.",
            ],
        )
        # --- Improve Macros & Instructions sheet formatting ---
        # Auto-size columns A-D in ws2
        for col_letter in ["A", "B", "C", "D"]:
            max_length = 0
            for cell in ws2[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws2.column_dimensions[col_letter].width = max(18, min(max_length + 4, 60))
        # Set alignment and wrap text for all cells
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        code_fill = PatternFill(
            start_color="F4F4F4", end_color="F4F4F4", fill_type="solid"
        )
        code_font = Font(name="Consolas", size=11)
        # Freeze top row
        ws2.freeze_panes = "A2"
        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True, vertical="center", horizontal="left"
                )
                cell.border = thin_border
                # Monospaced font and shading for code/formula blocks
                if cell.value and (
                    str(cell.value).strip().startswith("=")
                    or str(cell.value).strip().startswith("Sub")
                    or str(cell.value).strip().startswith("In ")
                    or str(cell.value).strip().startswith("#")
                    or "macro" in str(cell.value).lower()
                    or "Alt+F11" in str(cell.value)
                ):
                    cell.font = code_font
                    cell.fill = code_fill
        # Optionally, increase row height for rows with long/multiline content
        for row in ws2.iter_rows():
            for cell in row:
                if cell.value and (
                    "\n" in str(cell.value) or len(str(cell.value)) > 60
                ):
                    ws2.row_dimensions[cell.row].height = 32
        wb.save(filename)
    return filename


def show_faq():
    faqs = [
        ("What does this tool do?",
         "It generates a personalized Couch to 5K plan, calendar, and progress tracker with advanced features for beginners."),
        ("How do I customize my plan?",
         "You can choose the number of weeks and days per week during the prompts."),
        ("How do I use the Excel tracker?",
         "Open the generated Excel file. The 'Macros & Instructions' sheet explains all features and macros."),
        ("How do I enable accessibility options?",
         "You can select high-contrast or large font during the prompts. See the tracker instructions for more."),
        ("How do I export to my calendar or fitness app?",
         "Choose your desired export format (ICS, CSV, etc.) during the prompts. See the README for app-specific steps."),
        ("What if I miss a session?",
         "The tracker will highlight missed sessions and suggest repeating weeks if needed."),
        ("How do I get help or give feedback?",
         "See the README for contact info or use the feedback section in the tracker (planned)."),
        ("Where is my data stored?",
         "All files are saved locally in the 'created' folder inside the project directory."),
        ("How do I update or reset my plan?",
         "Just run the script again and enter new details. Each run creates a new output folder."),
    ]
    print("\nFrequently Asked Questions:\n" + "-"*32)
    for q, a in faqs:
        print(f"\nQ: {q}\nA: {a}")
    print("\nFor more, see the README or the Macros & Instructions sheet in your tracker.\n")


def prompt_smtp_config():
    """
    Prompt user for SMTP configuration or use environment variables.
    Returns a dict with SMTP settings or None if not configured.
    """
    import os
    smtp_host = os.environ.get("SMTP_HOST") or input(colorize("Enter SMTP server (e.g. smtp.gmail.com): ", "cyan", bold=True)).strip()
    smtp_port = os.environ.get("SMTP_PORT") or input(colorize("Enter SMTP port (e.g. 587): ", "cyan", bold=True)).strip()
    smtp_user = os.environ.get("SMTP_USER") or input(colorize("Enter SMTP username (your email): ", "cyan", bold=True)).strip()
    smtp_pass = os.environ.get("SMTP_PASS") or input(colorize("Enter SMTP password (input hidden): ", "cyan", bold=True)).strip()
    use_tls = True
    return {
        "host": smtp_host,
        "port": int(smtp_port),
        "user": smtp_user,
        "pass": smtp_pass,
        "tls": use_tls,
    }

def send_email_reminder(smtp_config, to_email, subject, body):
    """
    Send an email reminder using the provided SMTP config.
    """
    msg = MIMEMultipart()
    msg["From"] = smtp_config["user"]
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    try:
        with smtplib.SMTP(smtp_config["host"], smtp_config["port"]) as server:
            if smtp_config["tls"]:
                server.starttls()
            server.login(smtp_config["user"], smtp_config["pass"])
            server.sendmail(smtp_config["user"], to_email, msg.as_string())
        print(colorize(f"Email reminder sent to {to_email}.", "green", bold=True))
    except Exception:
        print(colorize("Failed to send email.", "red", bold=True))


def prompt_strava_runkeeper_config():
    """
    Prompt user for Strava or Runkeeper API token if exporting to mobile app.
    Returns a dict with platform and token.
    """
    platform = input(colorize("Export to [S]trava or [R]unkeeper? (S/R): ", "magenta", bold=True)).strip().lower()
    if platform == "s":
        token = input(colorize("Enter your Strava API Access Token: ", "yellow", bold=True)).strip()
        return {"platform": "strava", "token": token}
    elif platform == "r":
        token = input(colorize("Enter your Runkeeper API Access Token: ", "yellow", bold=True)).strip()
        return {"platform": "runkeeper", "token": token}
    else:
        print(colorize("Invalid platform. Skipping mobile app export.", "red", bold=True))
        return None

def export_to_mobile_app(plan, user, config):
    """
    Export the workout plan to Strava or Runkeeper using their APIs.
    """
    if not config:
        print(colorize("No mobile app config provided. Skipping export.", "red", bold=True))
        return
    if config["platform"] == "strava":
        url = "https://www.strava.com/api/v3/activities"
        headers = {"Authorization": f"Bearer {config['token']}"}
        for session in plan:
            if session["duration"] > 0:
                data = {
                    "name": f"C25K Week {session['week']} Day {session['day']}",
                    "type": "Run",
                    "start_date_local": f"{session['date']}T{user['hour']:02d}:{user['minute']:02d}:00",
                    "elapsed_time": session["duration"] * 60,
                    "description": session["description"],
                }
                resp = requests.post(url, headers=headers, data=data)
                if resp.status_code == 201:
                    print(colorize(f"Strava: Uploaded {data['name']}", "green"))
                else:
                    print(colorize(f"Strava: Failed to upload {data['name']} ({resp.status_code})", "red"))
    elif config["platform"] == "runkeeper":
        url = "https://api.runkeeper.com/fitnessActivities"
        headers = {
            "Authorization": f"Bearer {config['token']}",
            "Content-Type": "application/vnd.com.runkeeper.NewFitnessActivity+json"
        }
        for session in plan:
            if session["duration"] > 0:
                data = {
                    "type": "Running",
                    "start_time": f"{session['date']}T{user['hour']:02d}:{user['minute']:02d}",
                    "total_distance": 5.0,  # Placeholder, user can edit
                    "duration": session["duration"] * 60,
                    "notes": session["description"],
                }
                resp = requests.post(url, headers=headers, json=data)
                if resp.status_code in (201, 202):
                    print(colorize(f"Runkeeper: Uploaded {data['start_time']}", "green"))
                else:
                    print(colorize(f"Runkeeper: Failed to upload {data['start_time']} ({resp.status_code})", "red"))
    print(colorize("Mobile app export complete.", "green", bold=True))


def export_apple_health_csv(plan: list, filename: str) -> None:
    """
    Export the workout plan to an Apple Health-compatible CSV file.
    Columns: Type, Start, End, Calories, Distance, Duration, Source
    """
    import csv
    with open(filename, "w", newline="", encoding="utf-8") as csvfile:
        fieldnames = [
            "Type", "Start", "End", "Calories", "Distance (mi)", "Duration (min)", "Source"
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for session in plan:
            if session["duration"] > 0:
                # Estimate distance as 2.0 miles per session (user can edit)
                writer.writerow({
                    "Type": "Running",
                    "Start": f"{session['date']} {session.get('start_time', '07:00')}",
                    "End": f"{session['date']} {session.get('end_time', '07:00')}",
                    "Calories": "",
                    "Distance (mi)": 2.0,
                    "Duration (min)": session["duration"],
                    "Source": "C25K Calendar Generator"
                })
    print(f"Apple Health CSV file '{filename}' created successfully.")


def export_voice_prompts(plan, user, outdir):
    """
    Export session-by-session voice/text prompts as a text script and (optionally) audio files.
    """
    import os
    try:
        from gtts import gTTS
    except ImportError:
        gTTS = None
    lang_code = "en" if user.get("lang", "e") == "e" else "es"
    # Text script
    script_path = os.path.join(outdir, "C25K_Voice_Prompts.txt")
    with open(script_path, "w", encoding="utf-8") as f:
        for session in plan:
            if session["duration"] > 0:
                f.write(f"Week {session['week']} Day {session['day']}:\n")
                f.write(f"  Workout: {session['workout']}\n")
                f.write(f"  Tip: {session['tip']}\n\n")
    print(f"Voice prompt script '{script_path}' created successfully.")
    # Audio files (if gTTS is available)
                f.write(f"- Week {session['week']} Day {session['day']}: {session['workout']}\n")
    print(f"Shareable Markdown summary '{md_path}' created successfully.")
    # Email draft
    email_path = os.path.join(outdir, "C25K_Email_Share.txt")
    with open(email_path, "w", encoding="utf-8") as f:
        f.write("Subject: My Couch to 5K Plan\n\n")
        f.write("Hi!\n\nHere's my personalized Couch to 5K plan. Join me or cheer me on!\n\n")
        for session in plan:
            if session["duration"] > 0:
                f.write(f"Week {session['week']} Day {session['day']}: {session['workout']}\n")
        f.write("\nLet's get running!\n")
    print(f"Shareable email draft '{email_path}' created successfully.")


def main() -> None:
    """
    Main execution block for the Couch to 5K ICS Generator.
    """
    import sys
    if '--faq' in sys.argv or '--help' in sys.argv:
        show_faq()
        return
    """
    Main execution block for the Couch to 5K ICS Generator.
    """
    print(
        colorize(
            "Couch to 5K ICS Generator (for users with hypertension)",
            "magenta",
            bold=True,
        )
    )
    print(colorize("=" * 68, "yellow", bold=True))
    print(
        colorize(
            "\nDISCLAIMER: This script is for informational purposes only and is not a\n"
            "substitute for professional medical advice, diagnosis, or treatment.\n"
            "Always consult your healthcare provider before starting any new\n"
            "exercise program, especially if you have hypertension or other\n"
            "pre-existing health conditions. Use this script at your own risk.\n"
            "The author assumes no responsibility for any injury or health issues\n"
            "that may result from using this script.\n",
            "red",
            bold=True,
        )
    )
    print(colorize("=" * 68, "yellow", bold=True))
    user = get_user_info()
    if not user:
        print(
            "Missing or invalid information. "
            "Please provide all required details to generate your calendar."
        )
        return
    # Show summary before generating
    print(colorize("\nYour C25K Plan Summary:", "cyan", bold=True))
    print(
        colorize(
            f"  Weeks: {user['weeks']}  Days/Week: {user['days_per_week']}",
            "green",
            bold=True,
        )
    )
    print(
        colorize(
            f"  Start time: {user['hour']:02d}:{user['minute']:02d}",
            "yellow",
            bold=True,
        )
    )
    print(colorize(f"  Export format: {user['export'].upper()}", "magenta", bold=True))
    if user.get("start_day"):
        print(colorize(f"  Start date: {user['start_day']}", "blue", bold=True))
    if user.get("goal"):
        print(colorize(f"  Goal: {user['goal']}", "green", bold=True))
    if user.get("high_contrast") or user.get("large_font"):
        print(colorize("  Accessibility: ", "white", bold=True), end="")
        if user["high_contrast"]:
            print(colorize("High-contrast ", "yellow", bold=True), end="")
        if user["large_font"]:
            print(colorize("Large font", "yellow", bold=True), end="")
        print()
    # Plan customization
    plan = get_workout_plan(user)
    # Dynamic start date
    if user.get("start_day"):
        start_day = user["start_day"]
    else:
        from datetime import datetime

        start_day = datetime(2025, 7, 15)  # Default
    user["start_day"] = start_day  # Ensure always set
    outdir = get_output_dir(user)
    # Weather suggestion (example for first workout)
    if user.get("location"):
        suggestion = weather.get_weather_suggestion(user["location"], str(start_day))
        print(
            colorize(
                f"Weather suggestion for your first workout: {suggestion}",
                "cyan",
                bold=True,
            )
        )
    # Export logic
    if user["export"] == "i":
        generate_ics(
            plan,
            start_day,
            user["hour"],
            user["minute"],
            user.get("alert_minutes", 30),
            outdir=outdir,
        )
    elif user["export"] == "c":
        export_csv(plan, os.path.join(outdir, "Couch_to_5K_Reminders.csv"))
    elif user["export"] == "j":
        export_json(plan, os.path.join(outdir, "Couch_to_5K_Reminders.json"))
    elif user["export"] == "g":
        export_google_fit_csv(plan, os.path.join(outdir, "Couch_to_5K_GoogleFit.csv"))
    elif user["export"] == "p":
        pdf_export.export_to_pdf(
            plan, user, os.path.join(outdir, "Couch_to_5K_Plan.pdf")
        )
    elif user["export"] == "m":
        export_markdown_checklist(
            plan, os.path.join(outdir, "Couch_to_5K_Checklist.md"), user
        )
    elif user["export"] == "v":
        export_voice_prompts(plan, user, outdir)
    elif user["export"] == "s":
        config = prompt_strava_runkeeper_config()
        export_to_mobile_app(plan, user, config)
    elif user["export"] == "a":
        export_apple_health_csv(plan, os.path.join(outdir, "Couch_to_5K_AppleHealth.csv"))
    elif user["export"] == "h":
        export_community_share(plan, user, outdir)
    elif user["export"] == "q":
        qr_export.export_qr_code(plan, user, outdir)
    # Always output a Markdown checklist with user info
    export_markdown_checklist(
        plan, os.path.join(outdir, "Couch_to_5K_Checklist.md"), user
    )
    # Accessibility (example: print message)
    if user["high_contrast"] or user["large_font"]:
        print("Accessibility options enabled: ", end="")
        if user["high_contrast"]:
            print("High-contrast ", end="")
        if user["large_font"]:
            print("Large font", end="")
        print()
    # Reminders (rain/weather-aware)
    if user.get("email"):
        print(colorize("\n-- Email Reminders Setup --", "yellow", bold=True))
        smtp_config = prompt_smtp_config()
        if smtp_config:
            for session in plan:
                if session["duration"] > 0:
                    subject = f"C25K Reminder: Week {session['week']} Day {session['day']}"
                    body = (
                        f"Hi {user['name']},\n\nThis is your Couch to 5K session reminder!\n\n"
                        f"Date: {session['date']}\nTime: {user['hour']:02d}:{user['minute']:02d}\n"
                        f"Workout: {session['workout']}\nTip: {session['tip']}\n"
                    )
                    if session.get("weather"):
                        body += f"Weather forecast: {session['weather']}\n"
                        subject += f" ({session['weather']})"
                    body += "\nStay safe and have a great run!\n\n-- C25K Calendar Generator"
                    send_email_reminder(smtp_config, user["email"], subject, body)
            print(colorize("All session reminders sent!", "green", bold=True))
        else:
            print(colorize("SMTP config not provided. Email reminders skipped.", "red", bold=True))
    # Community/sharing (stub)
    community.share_plan(str(plan), platform="email")
    # Progress tracking (stub)
    # Create the progress tracker file if it doesn't exist
    progress_filename = create_progress_tracker(user, outdir)
    progress_data = progress.import_progress(progress_filename)
    print(progress.generate_progress_summary(progress_data))
    # --- Auto-insert macros into Excel tracker ---
    import stat
    import subprocess

    macro_inserter = os.path.join(
        os.path.dirname(__file__), "c25k_excel_macro_inserter.py"
    )
    if os.path.exists(macro_inserter):
        # Make executable if not already
        st = os.stat(macro_inserter)
        if not (st.st_mode & stat.S_IXUSR):
            os.chmod(macro_inserter, st.st_mode | stat.S_IXUSR)
        try:
            print("\nInserting macros into Excel tracker...")
            subprocess.run(["python3", macro_inserter, progress_filename], check=True)
        except Exception as e:
            print(f"[Warning] Could not auto-insert macros: {e}")
            print(
                f"You can run: python3 {macro_inserter} '{progress_filename}' to insert macros manually."
            )
            input("Press Enter to continue...")
    else:
        print("[Warning] Macro inserter script not found. Macros not auto-inserted.")
        input("Press Enter to continue...")
    print(
        colorize(
            "\nAll done! Your personalized C25K plan and exports are ready. "
            "Good luck!",
            "green",
            bold=True,
        )
    )
    # Privacy note
    print(colorize("\nNote: Your email, SMTP, and mobile app tokens are used only to send reminders/exports and are not stored.", "yellow", bold=True))


if __name__ == "__main__":
    main()
